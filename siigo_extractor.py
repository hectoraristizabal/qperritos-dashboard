"""
Qperritos - Extractor de ventas Siigo API
Extrae todas las facturas, limpia y transforma los datos,
y genera el archivo Excel listo para Power BI.

Uso: python siigo_extractor.py
Requisitos: pip install requests pandas openpyxl pytz python-dotenv

Credenciales: se leen SOLO de variables de entorno (SIIGO_USERNAME,
SIIGO_ACCESS_KEY, PARTNER_ID). En GitHub Actions vienen de los Secrets
del repo. Para correrlo en tu PC, creá un archivo ".env" en esta misma
carpeta (NUNCA lo subas a GitHub — agregalo a .gitignore) con:
    SIIGO_USERNAME=hector.aristizabal@eia.edu.co
    SIIGO_ACCESS_KEY=tu_access_key_aqui
    PARTNER_ID=PowerBI

VERSIÓN RÁPIDA — 2 cambios respecto a la anterior:
  1. Ventana incremental: en la primera corrida trae todo el histórico
     (desde FECHA_INICIO) como antes. De ahí en adelante solo vuelve a
     consultar los últimos DIAS_VENTANA días a Siigo — el resto del
     histórico ya transformado se reutiliza desde cache_facturas.json.
     Esto es lo que más pesa: con ~12.000 facturas acumuladas, antes se
     repaginaba TODO el histórico en cada corrida (100+ páginas); ahora
     solo se repagina la ventana reciente (1-3 páginas típicamente).
  2. Las consultas de detalle (necesarias solo cuando Siigo no trae la
     hora de creación en el listado) ahora se hacen en paralelo con un
     pool de 5 hilos en vez de una por una con time.sleep(0.4) entre
     cada una.
  Si algo se ve raro (huecos de datos, facturas editadas fuera de la
  ventana), borra cache_facturas.json para forzar una re-extracción
  completa del histórico — es el mismo comportamiento que el script
  tenía antes.
"""
import os
import json
import time
import requests
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

try:
    from dotenv import load_dotenv
    load_dotenv()  # carga variables desde un .env local si existe; en GitHub Actions no hace nada (no hay .env ahí, y no pasa nada si falta)
except ImportError:
    pass

CACHE_HORAS_PATH    = "cache_horas.json"
CACHE_FACTURAS_PATH = "cache_facturas.json"

# ─────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────
# Credenciales SOLO desde entorno (Secrets de GitHub Actions, o .env local).
# IMPORTANTE: nunca vuelvas a poner la clave real escrita aquí — este archivo
# es público en GitHub. Si alguna vez estuvo hardcodeada, esa clave ya quedó
# expuesta: hay que revocarla/regenerarla en Siigo, no basta con borrarla del código.
USERNAME   = os.environ.get("SIIGO_USERNAME")
ACCESS_KEY = os.environ.get("SIIGO_ACCESS_KEY")
PARTNER_ID = os.environ.get("PARTNER_ID", "PowerBI")

if not USERNAME or not ACCESS_KEY:
    raise SystemExit(
        "❌ Faltan credenciales de Siigo. Definí SIIGO_USERNAME y SIIGO_ACCESS_KEY "
        "como variables de entorno (Secrets en GitHub Actions, o un archivo .env local)."
    )

colombia = pytz.timezone("America/Bogota")
hoy = datetime.now(colombia)

FECHA_INICIO = "2026-05-01"
FECHA_FIN    = (hoy + timedelta(days=1)).strftime("%Y-%m-%d")

# Cuántos días hacia atrás se vuelven a consultar en cada corrida (adicional
# a lo que ya está en caché). Cubre facturas nuevas y correcciones recientes
# sin tener que repaginar todo el histórico. Súbelo si crees que se están
# quedando facturas por fuera; bájalo si quieres corridas aún más rápidas.
DIAS_VENTANA_INCREMENTAL = 10

MAX_HILOS_DETALLE = 5  # peticiones de detalle en paralelo (no subir mucho más: Siigo rate-limita)

print(f"📅 Rango total configurado: {FECHA_INICIO} hasta {FECHA_FIN}")


# ─────────────────────────────────────────
# 1. AUTENTICACIÓN
# ─────────────────────────────────────────
def get_token():
    r = requests.post(
        "https://api.siigo.com/auth",
        headers={"Content-Type": "application/json", "Partner-Id": PARTNER_ID},
        json={"username": USERNAME, "access_key": ACCESS_KEY},
        timeout=15
    )
    r.raise_for_status()
    print("✅ Token obtenido")
    return r.json()["access_token"]


# ─────────────────────────────────────────
# 2. EXTRACCIÓN — todas las facturas
# ─────────────────────────────────────────
def _get_con_reintentos(session, url, headers, params=None, timeout=15, intentos=6):
    """GET con reintentos y backoff exponencial ante 429 (Too Many Requests)
    o errores transitorios de red/servidor (5xx)."""
    espera = 2
    for intento in range(intentos):
        try:
            r = session.get(url, headers=headers, params=params, timeout=timeout)
        except requests.exceptions.RequestException as e:
            if intento == intentos - 1:
                raise
            print(f"    ⏳ Error de red ({e}) — reintentando en {espera:.0f}s")
            time.sleep(espera)
            espera = min(espera * 2, 60)
            continue

        if r.status_code == 429 or r.status_code >= 500:
            if intento == intentos - 1:
                r.raise_for_status()
            retry_after = r.headers.get("Retry-After")
            espera_real = float(retry_after) if retry_after else espera
            print(f"    ⏳ HTTP {r.status_code} — esperando {espera_real:.0f}s (intento {intento+1}/{intentos})")
            time.sleep(espera_real)
            espera = min(espera * 2, 60)
            continue

        r.raise_for_status()
        return r

    r.raise_for_status()
    return r


def _cargar_json(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return {}


def _guardar_json(path, data):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


# Sesión de requests por hilo — requests.Session no es seguro para compartir
# entre hilos, así que cada worker del pool de detalle usa la suya.
_local = threading.local()


def _sesion_hilo():
    if not hasattr(_local, "session"):
        _local.session = requests.Session()
    return _local.session


def _resolver_hora(factura_uid, factura_name, creado, headers):
    """Si falta la hora real (bug de Siigo: metadata.created vacío o
    '0001-01-01...'), consulta el detalle de esa factura puntual."""
    if len(creado) < 16 or creado.startswith("0001-01-01"):
        try:
            detalle = _get_con_reintentos(
                _sesion_hilo(),
                f"https://api.siigo.com/v1/invoices/{factura_uid}",
                headers
            ).json()
            creado = detalle.get("metadata", {}).get("created", "") or creado
        except Exception as e:
            print(f"  ⚠️  No se pudo recuperar hora real de {factura_name}: {e}")
    hora_str = creado[11:16] if len(creado) >= 16 else "00:00"
    return factura_uid, hora_str


def extraer_facturas(token):
    headers = {"Authorization": f"Bearer {token}", "Partner-Id": PARTNER_ID}
    session = requests.Session()

    # Caché de filas ya transformadas, por factura. Es lo que evita
    # repaginar el histórico completo en cada corrida: solo se vuelve a
    # pedir a Siigo la ventana reciente (DIAS_VENTANA_INCREMENTAL) y el
    # resto se toma de aquí tal cual.
    cache_facturas = _cargar_json(CACHE_FACTURAS_PATH)

    # Caché persistente factura_id -> "HH:MM", para no volver a pedir el
    # detalle de una factura cuya hora ya se resolvió alguna vez.
    cache_horas = _cargar_json(CACHE_HORAS_PATH)

    if cache_facturas:
        fecha_ventana = (hoy - timedelta(days=DIAS_VENTANA_INCREMENTAL)).strftime("%Y-%m-%d")
        fecha_inicio_consulta = max(FECHA_INICIO, fecha_ventana)
        print(f"📅 Caché existente ({len(cache_facturas)} facturas) — "
              f"consultando solo desde {fecha_inicio_consulta} (ventana incremental)")
    else:
        fecha_inicio_consulta = FECHA_INICIO
        print(f"📅 Sin caché — trayendo histórico completo desde {fecha_inicio_consulta} "
              f"(esta primera corrida sí tarda; las siguientes serán rápidas)")

    page = 1
    aciertos_cache = 0
    nuevas_consultas = 0
    facturas_en_ventana = []  # [(factura_uid, factura_dict, creado)]

    while True:
        r = _get_con_reintentos(
            session,
            "https://api.siigo.com/v1/invoices",
            headers,
            params={
                "date_start": fecha_inicio_consulta,
                "date_end":   FECHA_FIN,
                "page":          page,
                "page_size":     100
            }
        )
        data = r.json()
        resultados = data.get("results", [])

        if not resultados:
            break

        for f in resultados:
            creado = f.get("metadata", {}).get("created", "") or ""
            facturas_en_ventana.append((f["id"], f, creado))

        print(f"  Página {page} — {len(facturas_en_ventana)} facturas de la ventana recibidas")

        if not data.get("_links", {}).get("next"):
            break
        page += 1

    # Resolver horas: primero lo que ya está en caché (gratis), y en paralelo
    # solo lo que de verdad hace falta pedirle a Siigo.
    pendientes = []
    horas_resueltas = {}
    for factura_uid, f, creado in facturas_en_ventana:
        if factura_uid in cache_horas:
            horas_resueltas[factura_uid] = cache_horas[factura_uid]
            aciertos_cache += 1
        else:
            pendientes.append((factura_uid, f["name"], creado))

    if pendientes:
        print(f"  🔎 Resolviendo hora de {len(pendientes)} facturas nuevas "
              f"(hasta {MAX_HILOS_DETALLE} en paralelo)...")
        with ThreadPoolExecutor(max_workers=MAX_HILOS_DETALLE) as pool:
            futuros = {
                pool.submit(_resolver_hora, uid, nombre, creado, headers): uid
                for uid, nombre, creado in pendientes
            }
            for fut in as_completed(futuros):
                factura_uid, hora_str = fut.result()
                horas_resueltas[factura_uid] = hora_str
                cache_horas[factura_uid] = hora_str
                nuevas_consultas += 1

    _guardar_json(CACHE_HORAS_PATH, cache_horas)

    # Reconstruir/actualizar el caché de filas por factura con lo que vino
    # en esta ventana (esto también recoge ediciones recientes en Siigo).
    for factura_uid, f, _creado in facturas_en_ventana:
        hora_str = horas_resueltas[factura_uid]
        filas = []
        for item in f.get("items", []):
            filas.append({
                "factura_id":  f["name"],
                "fecha":       f["date"],
                "hora":        hora_str,
                "hora_Militar": hora_str[:2],
                "forma_pago":  f["payments"][0]["name"] if f.get("payments") else "",
                "producto_cod":  item["code"],
                "producto":      item["description"],
                "cantidad":      item["quantity"],
                "precio_unit":   item["price"],
                "total_item":    item["total"],
                "bodega_id":     item.get("warehouse", {}).get("id", ""),
                "bodega_nombre": item.get("warehouse", {}).get("name", ""),
            })
        cache_facturas[factura_uid] = filas

    _guardar_json(CACHE_FACTURAS_PATH, cache_facturas)

    todas = [fila for filas in cache_facturas.values() for fila in filas]

    print(f"✅ Extracción completa: {len(todas)} registros ({len(cache_facturas)} facturas en total)")
    print(f"   Facturas de la ventana consultada: {len(facturas_en_ventana)} "
          f"(horas desde caché: {aciertos_cache} | horas consultadas a la API: {nuevas_consultas})")
    return pd.DataFrame(todas)


# ─────────────────────────────────────────
# 3. LIMPIEZA Y TRANSFORMACIÓN
# ─────────────────────────────────────────
def transformar(df):


    # Acortar nombres de productos
    df["producto"] = df["producto"].replace({
        "Coca Cola Original": "Coca Original",
        "Coca Cola Zero":     "Coca Zero",
    })

    # Reemplazar hora_Militar con franja de 30 minutos
    df["hora_Militar"] = df["hora"].apply(
        lambda h: f"{h[:2]}:00" if int(h[3:5]) < 30 else f"{h[:2]}:30"
    )

    # Todas las saborizadas — detecta cualquier nombre que contenga "Saborizada"
    df["producto"] = df["producto"].apply(
        lambda x: x.replace("Saborizada ", "Sab. ") if "Saborizada" in str(x) else x
    )
    # Factura_E: True si es factura electrónica (empieza con FV-3)
    df["factura_E"] = df["factura_id"].str.startswith("FV-3")

    # Acortar nombres de formas de pago
    df["forma_pago"] = df["forma_pago"].replace({
        "Qr Banco Bogotá":        "QR Bogotá",
        "Tarjeta Débito":         "T. Débito",
        "Tarjeta Crédito":        "T. Crédito",
    })


    # precio_base: precio unitario sin impuesto
    # Si es FV-3 (electrónica): total_item / cantidad / 1.08
    # Si es FV-2 (no electrónica): total_item / cantidad
    df["precio_base"] = df["total_item"] / df["cantidad"]
    df["precio_base"] = df.apply(
        lambda r: round(r["precio_base"] / 1.08, 2) if r["factura_E"]
                  else round(r["precio_base"], 2),
        axis=1
    )

    # impuesto: 8% del precio_base solo en facturas electrónicas
    df["impuesto"] = df.apply(
        lambda r: round(r["precio_base"] * 0.08, 2) if r["factura_E"] else 0.0,
        axis=1
    )

    # total_venta: venta sin impuesto
    df["total_venta"] = (df["cantidad"] * df["precio_base"]).round(2)

    # total_impuestos: impuesto total del ítem
    df["total_impuestos"] = (df["cantidad"] * df["impuesto"]).round(2)

    # Eliminar columnas intermedias ya no necesarias
    df = df.drop(columns=["precio_unit", "total_item"])

    # Columnas de fecha
    fecha = pd.to_datetime(df["fecha"])
    df["año"]            = fecha.dt.year
    df["mes"]            = fecha.dt.month
    df["dia"]            = fecha.dt.day
    dias_es = {
        "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
        "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"
    }

    # Primero crear dia_semana
    dias_es = {
        "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
        "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"
    }
    df["dia_semana"] = fecha.dt.day_name().map(dias_es)

    # DESPUÉS crear orden_dia
    orden_dias = {
        "Lunes": 1, "Martes": 2, "Miércoles": 3,
        "Jueves": 4, "Viernes": 5, "Sábado": 6, "Domingo": 7
    }
    df["orden_dia"] = df["dia_semana"].map(orden_dias)

    # Reordenar columnas
    df = df[[
        "factura_id", "año", "mes", "dia", "dia_semana", "orden_dia",
        "fecha", "hora", "hora_Militar", "bodega_id", "bodega_nombre",
        "forma_pago", "producto_cod", "producto", "cantidad",
        "factura_E", "precio_base", "impuesto", "total_venta", "total_impuestos"
    ]]

    print(f"✅ Transformación completa")
    print(f"   Facturas electrónicas (FV-3): {df[df['factura_E']]['factura_id'].nunique()}")
    print(f"   Facturas normales     (FV-2): {df[~df['factura_E']]['factura_id'].nunique()}")
    print(f"   Total ventas (sin impuesto):  ${df['total_venta'].sum():,.0f}")
    print(f"   Total impuestos:              ${df['total_impuestos'].sum():,.0f}")

    print(f"Fecha máxima en datos: {df['fecha'].max()}")
    print(f"Fecha mínima en datos: {df['fecha'].min()}")
    print(f"Facturas de hoy: {len(df[df['fecha'] == FECHA_FIN])}")
    return df


# ─────────────────────────────────────────
# 4. GUARDAR EXCEL
# ─────────────────────────────────────────
def guardar_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    thin     = Side(style="thin", color="BFBFBF")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols_moneda = ["precio_base","impuesto","total_venta","total_impuestos"]

    # Encabezados
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.upper().replace("_", " "))
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    # Datos
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = Font(name="Arial", size=10)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            col_name = df.columns[col_idx - 1]
            if col_name in cols_moneda:
                cell.number_format = "#,##0.00"
            elif col_name == "cantidad":
                cell.number_format = "#,##0"

    # Ancho columnas y filtros
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(len(col), df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 28)
    ws.freeze_panes  = "A2"
    ws.auto_filter.ref = ws.dimensions

    archivo = "ventas_qperritos.xlsx"
    wb.save(archivo)
    print(f"✅ Archivo guardado: {archivo}")
    print(f"   Filas: {len(df)} | Columnas: {len(df.columns)}")
    return archivo


# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
if __name__ == "__main__":
    _inicio = time.time()
    print("=" * 50)
    print("  QPERRITOS — Extractor Siigo API")
    print(f"  Período total: {FECHA_INICIO} → {FECHA_FIN}")
    print("=" * 50)

    token = get_token()
    df    = extraer_facturas(token)
    df    = transformar(df)
    guardar_excel(df)

    print(f"\n✅ Proceso completado exitosamente en {time.time() - _inicio:.1f}s")
